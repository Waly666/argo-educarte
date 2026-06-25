#!/usr/bin/env python3
"""
Genera un Excel compatible con ARGO Migración desde Microsoft Access.

Uso:
  python scripts/access-to-argo-xlsx.py "C:\\datos\\app.accdb" access-migracion.mapping.json salida.xlsx

Requisitos:
  pip install pyodbc openpyxl
"""

import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import pyodbc
    from openpyxl import Workbook
except ImportError:
    print("Instale dependencias: pip install pyodbc openpyxl")
    sys.exit(1)

HOJAS = {
    "programas": "Programas",
    "alumnos": "Alumnos",
    "matriculas": "Matriculas",
    "pagos": "Pagos",
    "certificados": "Certificados",
}

COLUMNAS_ORDEN = {
    "programas": [
        "codigoPrograma", "nombrePrograma", "tipoCapacitacion", "horas", "semestres",
        "diasVencimiento", "tarifa1", "tarifa2", "tarifa3", "tarifaVirtual",
    ],
    "alumnos": [
        "numDoc", "tipoDoc", "nombre1", "nombre2", "apellido1", "apellido2",
        "fechaNacimiento", "genero", "celular", "correo", "direccion", "municipio", "observaciones",
    ],
    "matriculas": [
        "numDoc", "codigoPrograma", "fechaMatricula", "valorTotal", "valorPagado", "estado", "observaciones",
    ],
    "pagos": [
        "numDoc", "numeroRecibo", "fecha", "valor", "formaPago", "concepto", "observaciones",
    ],
    "certificados": [
        "numDoc", "nombreTitular", "codVerificacion", "codigoPrograma", "codigoCertificado", "nombreCurso", "horas",
        "fechaEmision", "fechaVencimiento", "numActa", "numFolio", "numRunt", "estado",
    ],
}


def limpiar_num_doc(v):
    if v is None:
        return ""
    s = re.sub(r"\D", "", str(v).strip())
    return s if s else ""


def fmt_valor(v):
    if v is None or v == "":
        return ""
    if isinstance(v, (int, float)):
        return v
    try:
        return float(str(v).replace(",", "").replace("$", "").strip())
    except ValueError:
        return str(v)


def fmt_fecha(v):
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.isoformat()
    return str(v).strip()


def leer_tabla(conn, tabla, map_col):
    cur = conn.cursor()
    cur.execute(f"SELECT * FROM [{tabla}]")
    cols_access = [d[0] for d in cur.description]
    inv = {access: argo for argo, access in map_col.items() if access in cols_access}
    filas = []
    for row in cur.fetchall():
        d = dict(zip(cols_access, row))
        out = {}
        for argo, access in map_col.items():
            val = d.get(access)
            if argo == "numDoc":
                out[argo] = limpiar_num_doc(val)
            elif argo in ("valorTotal", "valorPagado", "valor", "horas", "tarifa1", "tarifa2", "tarifa3", "tarifaVirtual", "semestres", "diasVencimiento"):
                out[argo] = fmt_valor(val)
            elif "fecha" in argo.lower() or argo == "fecha":
                out[argo] = fmt_fecha(val)
            else:
                out[argo] = "" if val is None else str(val).strip()
        if out.get("numDoc") or any(out.get(k) for k in out if k != "numDoc"):
            filas.append(out)
    return filas


def main():
    if len(sys.argv) < 4:
        print(__doc__)
        sys.exit(1)

    accdb = sys.argv[1]
    mapping_path = Path(sys.argv[2])
    salida = Path(sys.argv[3])

    mapping = json.loads(mapping_path.read_text(encoding="utf-8"))
    conn_str = (
        r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};"
        rf"DBQ={accdb};"
    )
    conn = pyodbc.connect(conn_str)

    wb = Workbook()
    wb.remove(wb.active)

    ws_inst = wb.create_sheet("Instrucciones", 0)
    ws_inst["A1"] = "Generado desde Access — revise filas antes de importar en ARGO → Sistema → Migración"
    ws_inst["A2"] = f"Origen: {accdb}"
    ws_inst["A3"] = f"Mapeo: {mapping_path.name}"

    for clave, titulo_hoja in HOJAS.items():
        cfg = mapping.get(clave)
        if not cfg or not cfg.get("tabla"):
            continue
        filas = leer_tabla(conn, cfg["tabla"], cfg.get("columnas", {}))
        headers = COLUMNAS_ORDEN[clave]
        ws = wb.create_sheet(titulo_hoja)
        ws.append(headers)
        for f in filas:
            ws.append([f.get(h, "") for h in headers])
        print(f"{titulo_hoja}: {len(filas)} filas")

    conn.close()
    wb.save(salida)
    print(f"Guardado: {salida.resolve()}")


if __name__ == "__main__":
    main()
